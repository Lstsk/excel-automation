#!/usr/bin/env python3
"""
Excel Processing Engine
Handles reading Excel templates and inserting parsed shipment data.
"""

import os
import shutil
from datetime import datetime
from typing import List, Dict, Optional, Tuple
import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
import re

from llm_parser import ShipmentData
from config import get_config

class ExcelProcessor:
    """Processes Excel templates and inserts shipment data"""
    
    def __init__(self, template_path: Optional[str] = None):
        """
        Initialize the Excel processor
        
        Args:
            template_path: Path to Excel template file
        """
        self.config = get_config()
        self.template_path = template_path or self.config["excel"]["template_file"]
        self.sheet_name = self.config["excel"]["sheet_name"]
        self.columns = self.config["excel"]["columns"]
        self.field_mapping = self.config["field_mapping"]
        self.defaults = self.config["defaults"]
        
        # Validate template exists
        if not os.path.exists(self.template_path):
            raise FileNotFoundError(f"Template file not found: {self.template_path}")
    
    def load_template(self) -> Tuple[Workbook, any]:
        """
        Load the Excel template
        
        Returns:
            Tuple of (workbook, worksheet)
        """
        try:
            wb = load_workbook(self.template_path)
            
            # Try to get the specified sheet, fallback to active sheet
            if self.sheet_name in wb.sheetnames:
                ws = wb[self.sheet_name]
            else:
                ws = wb.active
                print(f"Warning: Sheet '{self.sheet_name}' not found, using active sheet: {ws.title}")
            
            return wb, ws
        except Exception as e:
            raise Exception(f"Error loading template: {e}")
    
    def find_next_data_row(self, ws) -> int:
        """
        Find the next available row for data insertion
        
        Args:
            ws: Worksheet object
            
        Returns:
            Row number for next data insertion
        """
        data_start_row = self.config["excel"]["data_start_row"]
        
        # Start from the data start row and find the first empty row
        for row in range(data_start_row, ws.max_row + 2):
            # Check if row is empty (specifically check key columns)
            key_columns = ["A", "C", "F"]  # Case number, product name, price
            is_empty = True
            
            for col_letter in key_columns:
                cell_value = ws[f"{col_letter}{row}"].value
                if cell_value is not None and str(cell_value).strip():
                    is_empty = False
                    break
            
            if is_empty:
                return row
        
        # If no empty row found, return next row after max_row
        return ws.max_row + 1
    
    def generate_case_number(self, ws, row_num: int) -> str:
        """
        Generate case number for the shipment
        
        Args:
            ws: Worksheet object
            row_num: Current row number
            
        Returns:
            Case number string (e.g., "Case 1", "Case 2")
        """
        data_start_row = self.config["excel"]["data_start_row"]
        case_num = row_num - data_start_row + 1
        return f"Case {case_num}"
    
    def clean_price(self, price_str: str) -> str:
        """
        Clean and format price string
        
        Args:
            price_str: Raw price string (e.g., "30$", "25美金")
            
        Returns:
            Cleaned numeric price string
        """
        if not price_str:
            return ""
        
        # Extract numeric value
        price_match = re.search(r'(\d+(?:\.\d+)?)', price_str)
        if price_match:
            return price_match.group(1)
        
        return ""
    
    def normalize_courier(self, courier: str) -> str:
        """
        Normalize courier company name
        
        Args:
            courier: Raw courier name
            
        Returns:
            Normalized courier name
        """
        if not courier:
            return ""
        
        courier_mappings = self.config["courier_mappings"]
        
        # Check if it's in the mapping
        for key, value in courier_mappings.items():
            if key in courier:
                return value
        
        # If not found, return as is (might already be normalized)
        return courier
    
    def format_receipt_date(self, date_str: str) -> str:
        """
        Format receipt date to MM/DD format

        Args:
            date_str: Date string in YYYY-MM-DD format

        Returns:
            Formatted date string in MM/DD format
        """
        if not date_str:
            return ""

        try:
            # Parse YYYY-MM-DD format
            if len(date_str) == 10 and date_str.count('-') == 2:
                year, month, day = date_str.split('-')
                return f"{month.lstrip('0')}/{day.lstrip('0')}"
            else:
                return date_str  # Return as-is if not in expected format
        except:
            return date_str

    def insert_shipment_data(self, ws, shipment: ShipmentData, row_num: int):
        """
        Insert a single shipment data into the worksheet

        Args:
            ws: Worksheet object
            shipment: ShipmentData object
            row_num: Row number to insert data
        """
        # Generate case number
        case_number = self.generate_case_number(ws, row_num)
        ws[f"A{row_num}"] = case_number

        # Package unit (leave empty as requested)
        ws[f"B{row_num}"] = self.defaults["package_unit"]

        # Chinese product name
        ws[f"C{row_num}"] = shipment.货物名称

        # English product name (if provided)
        ws[f"D{row_num}"] = shipment.英文品名 or self.defaults["english_name"]

        # Quantity/specification (use the complete quantity field)
        ws[f"E{row_num}"] = shipment.数量 or ""

        # Unit price (cleaned and formatted as number)
        cleaned_price = self.clean_price(shipment.单价)
        if cleaned_price:
            try:
                # Convert to float for proper Excel formatting
                ws[f"F{row_num}"] = float(cleaned_price)
            except ValueError:
                ws[f"F{row_num}"] = cleaned_price

        # Total price (formula or same as unit price)
        if cleaned_price:
            # Use Excel formula for total price
            ws[f"G{row_num}"] = f"=F{row_num}"

        # Volume (optional, default empty)
        ws[f"H{row_num}"] = self.defaults["volume"]

        # Weight (optional, default empty)
        ws[f"I{row_num}"] = self.defaults["weight"]

        # Courier company (normalized)
        normalized_courier = self.normalize_courier(shipment.快递公司)
        ws[f"J{row_num}"] = normalized_courier

        # Courier number (tracking number) - format as text to preserve leading zeros
        if shipment.快递单号:
            # Ensure tracking number is treated as text
            ws[f"K{row_num}"] = str(shipment.快递单号)

        # Receipt date (formatted as MM/DD)
        formatted_date = self.format_receipt_date(shipment.入仓日期)
        ws[f"L{row_num}"] = formatted_date
    
    def process_shipments(self, shipments: List[ShipmentData], output_path: Optional[str] = None) -> str:
        """
        Process multiple shipments and save to Excel file
        
        Args:
            shipments: List of ShipmentData objects
            output_path: Output file path (if None, auto-generate)
            
        Returns:
            Path to the saved Excel file
        """
        if not shipments:
            raise ValueError("No shipments to process")
        
        # Load template
        wb, ws = self.load_template()
        
        # Find starting row for data insertion
        start_row = self.find_next_data_row(ws)
        
        # Insert each shipment
        for i, shipment in enumerate(shipments):
            row_num = start_row + i
            self.insert_shipment_data(ws, shipment, row_num)
        
        # Generate output path if not provided
        if not output_path:
            timestamp = datetime.now().strftime(self.config["output"]["timestamp_format"])
            filename = self.config["output"]["filename_format"].format(timestamp=timestamp)
            output_path = os.path.join(self.config["output"]["output_dir"], filename)
        
        # Ensure output directory exists
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        
        # Save the workbook
        wb.save(output_path)
        
        print(f"Successfully processed {len(shipments)} shipments")
        print(f"Output saved to: {output_path}")
        
        return output_path
    
    def create_backup(self, source_path: str) -> str:
        """
        Create a backup of the source file
        
        Args:
            source_path: Path to source file
            
        Returns:
            Path to backup file
        """
        backup_dir = self.config["output"]["backup_dir"]
        os.makedirs(backup_dir, exist_ok=True)
        
        timestamp = datetime.now().strftime(self.config["output"]["timestamp_format"])
        filename = os.path.basename(source_path)
        name, ext = os.path.splitext(filename)
        backup_filename = f"{name}_backup_{timestamp}{ext}"
        backup_path = os.path.join(backup_dir, backup_filename)
        
        shutil.copy2(source_path, backup_path)
        return backup_path
    
    def analyze_missing_fields(self, shipment: ShipmentData) -> Dict[str, str]:
        """
        Analyze what fields are missing and provide suggestions

        Args:
            shipment: ShipmentData object

        Returns:
            Dictionary of missing fields and suggestions
        """
        missing = {}

        # Check each field and provide suggestions
        if not shipment.货物名称:
            missing["货物名称"] = "Product name is required"

        if not shipment.单价:
            missing["单价"] = "Price is recommended for customs declaration"

        if not shipment.快递公司:
            missing["快递公司"] = "Courier company helps with tracking"

        if not shipment.快递单号:
            missing["快递单号"] = "Tracking number enables shipment monitoring"

        if not shipment.入仓日期:
            missing["入仓日期"] = "Receipt date helps with inventory management"

        if not shipment.数量:
            missing["数量"] = "Quantity specification recommended"

        if not shipment.英文品名:
            missing["英文品名"] = "English name may be required for customs"

        return missing

    def auto_complete_shipment(self, shipment: ShipmentData) -> ShipmentData:
        """
        Auto-complete missing fields with intelligent defaults

        Args:
            shipment: Original ShipmentData object

        Returns:
            Enhanced ShipmentData object with auto-completed fields
        """
        # Create a copy to avoid modifying original
        enhanced = ShipmentData(
            货物名称=shipment.货物名称,
            数量=shipment.数量,
            单价=shipment.单价,
            快递公司=shipment.快递公司,
            快递单号=shipment.快递单号,
            入仓日期=shipment.入仓日期,
            英文品名=shipment.英文品名
        )

        # Auto-complete quantity if missing but can be inferred
        if not enhanced.数量 and enhanced.货物名称:
            # Look for quantity indicators in product name
            import re
            qty_patterns = [
                r'(\d+(?:托|箱|个|件|张|套|台|只|条|包))',
                r'(一|二|三|四|五|六|七|八|九|十)(托|箱|个|件|张|套|台|只|条|包)',
            ]
            for pattern in qty_patterns:
                match = re.search(pattern, enhanced.货物名称)
                if match:
                    if len(match.groups()) == 1:
                        enhanced.数量 = match.group(1)
                    else:
                        # Convert Chinese numbers to Arabic and combine with unit
                        chinese_to_arabic = {'一': '1', '二': '2', '三': '3', '四': '4', '五': '5',
                                           '六': '6', '七': '7', '八': '8', '九': '9', '十': '10'}
                        num = chinese_to_arabic.get(match.group(1), match.group(1))
                        enhanced.数量 = f"{num}{match.group(2)}"
                    break

        # Auto-complete English name with common translations
        if not enhanced.英文品名 and enhanced.货物名称:
            common_translations = {
                "地板": "Flooring",
                "按摩床": "Massage Table",
                "折叠按摩床": "Folding Massage Table",
                "电子产品": "Electronic Products",
                "家具": "Furniture",
                "服装": "Clothing",
                "玩具": "Toys",
                "工具": "Tools",
                "厨具": "Kitchenware",
                "装饰品": "Decorations"
            }

            for chinese, english in common_translations.items():
                if chinese in enhanced.货物名称:
                    enhanced.英文品名 = english
                    break

        # Auto-complete receipt date with today if missing
        if not enhanced.入仓日期:
            from datetime import datetime
            enhanced.入仓日期 = datetime.now().strftime("%Y-%m-%d")

        return enhanced

    def validate_shipment_data(self, shipment: ShipmentData) -> List[str]:
        """
        Validate shipment data

        Args:
            shipment: ShipmentData object

        Returns:
            List of validation error messages
        """
        errors = []
        validation_rules = self.config["validation"]

        # Check required fields
        for field in validation_rules["required_fields"]:
            # Handle field name mapping
            attr_name = field.replace("/", "_").replace("货物名称", "货物名称")
            value = getattr(shipment, attr_name, "")
            if not value or not str(value).strip():
                errors.append(f"Required field missing: {field}")

        # Validate price format
        if shipment.单价:
            cleaned_price = self.clean_price(shipment.单价)
            if not re.match(validation_rules["price_pattern"], cleaned_price):
                errors.append(f"Invalid price format: {shipment.单价}")

        # Validate product name length
        if len(shipment.货物名称) > validation_rules["max_product_name_length"]:
            errors.append(f"Product name too long: {len(shipment.货物名称)} characters")

        return errors

def test_excel_processor():
    """Test the Excel processor"""
    print("Testing Excel Processor...")
    
    # Create sample shipment data
    shipments = [
        ShipmentData(
            货物名称="地板",
            数量_单位="1托",
            单价="30$",
            快递公司="中通",
            快递单号="202242834846",
            入仓日期="2025-07-05"
        ),
        ShipmentData(
            货物名称="折叠按摩床",
            数量_单位="1张",
            单价="25$",
            快递公司="中通快递",
            快递单号="76018395245100010001",
            入仓日期="2025-07-04"
        )
    ]
    
    try:
        processor = ExcelProcessor()
        
        # Validate shipments
        for i, shipment in enumerate(shipments):
            errors = processor.validate_shipment_data(shipment)
            if errors:
                print(f"Validation errors for shipment {i+1}: {errors}")
            else:
                print(f"Shipment {i+1} validation: ✓ Passed")
        
        # Process shipments
        output_path = processor.process_shipments(shipments)
        print(f"Test completed successfully: {output_path}")
        
    except Exception as e:
        print(f"Test failed: {e}")

if __name__ == "__main__":
    test_excel_processor()
