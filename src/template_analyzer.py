#!/usr/bin/env python3
"""
Excel Template Analyzer
Analyzes the structure of the Excel template to understand column mappings and data insertion points.
"""

import pandas as pd
import openpyxl
from openpyxl import load_workbook
import sys
import os

def analyze_excel_template(file_path):
    """
    Analyze the Excel template structure
    """
    print(f"Analyzing Excel template: {file_path}")
    print("=" * 50)
    
    try:
        # Load workbook with openpyxl to get detailed structure
        wb = load_workbook(file_path, data_only=False)
        print(f"Workbook sheets: {wb.sheetnames}")
        
        for sheet_name in wb.sheetnames:
            print(f"\n--- Sheet: {sheet_name} ---")
            ws = wb[sheet_name]
            
            # Get dimensions
            print(f"Max row: {ws.max_row}")
            print(f"Max column: {ws.max_column}")
            
            # Read first 10 rows to understand structure
            print("\nFirst 10 rows:")
            for row_num in range(1, min(11, ws.max_row + 1)):
                row_data = []
                for col_num in range(1, min(11, ws.max_column + 1)):  # First 10 columns
                    cell = ws.cell(row=row_num, column=col_num)
                    value = cell.value
                    if value is not None:
                        row_data.append(str(value)[:20])  # Truncate long values
                    else:
                        row_data.append("")
                print(f"Row {row_num}: {row_data}")
            
            # Look for header row (likely contains field names)
            print("\nLooking for header patterns...")
            for row_num in range(1, min(6, ws.max_row + 1)):
                row_values = []
                for col_num in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    if cell.value:
                        row_values.append(str(cell.value))
                
                if row_values:
                    print(f"Row {row_num} headers: {row_values}")
        
        # Also try reading with pandas to see data structure
        print(f"\n--- Pandas Analysis ---")
        try:
            df = pd.read_excel(file_path, sheet_name=None)  # Read all sheets
            for sheet_name, data in df.items():
                print(f"\nSheet '{sheet_name}' shape: {data.shape}")
                print("Columns:", list(data.columns))
                print("First few rows:")
                print(data.head())
        except Exception as e:
            print(f"Pandas reading error: {e}")
            
    except Exception as e:
        print(f"Error analyzing template: {e}")
        return None

def identify_data_insertion_points(file_path):
    """
    Identify where new data should be inserted
    """
    print(f"\n--- Data Insertion Analysis ---")
    
    try:
        wb = load_workbook(file_path)
        ws = wb.active  # Use the active sheet
        
        # Find the last row with data
        last_row = 0
        for row in range(1, ws.max_row + 1):
            has_data = False
            for col in range(1, ws.max_column + 1):
                if ws.cell(row=row, column=col).value is not None:
                    has_data = True
                    break
            if has_data:
                last_row = row
        
        print(f"Last row with data: {last_row}")
        print(f"New data should be inserted starting from row: {last_row + 1}")
        
        # Identify column structure around the data area
        if last_row > 0:
            print(f"\nData structure around row {last_row}:")
            for col in range(1, min(11, ws.max_column + 1)):
                cell_value = ws.cell(row=last_row, column=col).value
                col_letter = openpyxl.utils.get_column_letter(col)
                print(f"Column {col_letter} ({col}): {cell_value}")
                
    except Exception as e:
        print(f"Error in insertion point analysis: {e}")

if __name__ == "__main__":
    # Template file path
    template_path = "环亚客户 LeShuiju自行申报货物表- - 2y - Copy.xlsx"
    
    if os.path.exists(template_path):
        analyze_excel_template(template_path)
        identify_data_insertion_points(template_path)
    else:
        print(f"Template file not found: {template_path}")
        print("Available files:")
        for file in os.listdir("."):
            if file.endswith(('.xlsx', '.xls')):
                print(f"  - {file}")
